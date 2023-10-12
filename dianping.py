from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
import pandas as pd
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import urllib.request
from urllib.error import URLError
import os, os.path
import re
from datetime import datetime
#import win32com.client 
import openpyxl as xl
from openpyxl.formula.translate import Translator
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
import smtplib
from email import encoders
from selenium.webdriver.common.action_chains import ActionChains
#import pyautogui
from selenium.webdriver.chrome.options import Options

from baidutranslate import *



chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('--user-agent="Mozilla/5.0 (iPhone; CPU iPhone OS 15_5 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) CriOS/102.0.5005.87 Mobile/15E148 Safari/604.1"')

webdriver = webdriver.Chrome(chrome_options=chrome_options)

today = datetime.today().strftime('%Y%m%d')
excelname = 'dianping.xlsx'
try:
    os.remove(excelname)
except:
    pass
wb = xl.Workbook()

wb.active.title = 'Dianping'   
wb.create_sheet(title = 'Dianpingfood')
wb.create_sheet(title = 'Dianpingtag') 
wb.create_sheet(title = 'Dianpingrank') 

wsd = wb['Dianping']
wsf = wb['Dianpingfood']
wst = wb['Dianpingtag']
wsr = wb['Dianpingrank']
dp = 1
df = 1
dt = 1
dr = 1
i =0

x = ['http://www.dianping.com/shop/H42o9y1W7EvMw3jg',
'http://www.dianping.com/shop/H9Jb2XH5MBlWZcGs',
'http://www.dianping.com/shop/G2A1kkW8qcyKw3tg',
'https://www.dianping.com/shopshare/E3m6u0VARLddOis7?msource=Appshare2021&utm_source=shop_share',
'http://www.dianping.com/shop/H6Aeq6ALX5NcxLnb',
'http://www.dianping.com/shop/k3KzmcTeuWbtg7Qy',
'http://www.dianping.com/shop/G9sdPLgfWhaEnAqm',
'http://www.dianping.com/shop/k4dsysAk5NMq0fUl',
'http://www.dianping.com/shop/G7gCkATdqC1e4AIN',
'http://www.dianping.com/shop/H9aSMwelq3UwiaoU',
'http://www.dianping.com/shop/k4f4JnTPYZNw9v3m',
'http://www.dianping.com/shop/HalnrWoNkOVXAg2v',
'http://www.dianping.com/shop/k1GoqrPCbL1RSuuL',
'http://www.dianping.com/shop/l2rBlWuR3I3G37ZD',
' https://www.dianping.com/shopshare/Ga71vqJoQzb65huX?msource=Appshare2021&utm_source=shop_share',
'http://www.dianping.com/shop/k14gkQMxQ4Av3NA1',
'http://www.dianping.com/shop/l1kTOZw7jKRcLBGl',
'http://www.dianping.com/shop/G1TuGsq7Z0BLqqtV',
'http://www.dianping.com/shop/H21jKyqUHiuqcAzj',
'http://www.dianping.com/shop/H2kFyBG3jVFyxXGo',
'http://www.dianping.com/shop/G4bweBe9RaKEvWAw',
'http://www.dianping.com/shop/HaPWqTTOypkspbhZ',
'http://www.dianping.com/shop/kaFQV1F3oXObFh0U',
'http://www.dianping.com/shop/G7zTrLCf7Xtm15xH',
'http://www.dianping.com/shop/H5eCnLwNQ6G5jScB',
'http://www.dianping.com/shop/H4rXc0LAmAMNOKPk',
'http://www.dianping.com/shop/GaKv5NlwdzrxVyX1',
'http://www.dianping.com/shop/G82jVKf2ttihHY0J',
'http://www.dianping.com/shop/H3Avv12Mib11uidi',
'http://www.dianping.com/shop/iARvBEMNh2O7D2ub',
'http://www.dianping.com/shop/l4ahEG1ux4tdp4ks',
'http://www.dianping.com/shop/l8nI32cAvOmjxUex',
'https://www.dianping.com/shopshare/H8FvBnpiGVkDFur3?msource=Appshare2021&utm_source=shop_share',
'https://www.dianping.com/shopshare/Ga71vqJoQzb65huX',
'https://www.dianping.com/shopshare/H8kUL2Du3RlY5Rj5?msource=Appshare2021&utm_source=shop_share',
'https://www.dianping.com/shopshare/G5uCuVpOz3xfekGc?msource=Appshare2021&utm_source=shop_share',
'https://www.dianping.com/shopshare/G1tA3WT9H2BAta4Z?msource=Appshare2021&utm_source=shop_share',
'https://www.dianping.com/shopshare/l9hMHgJrY6hE3Ii3?msource=Appshare2021&utm_source=shop_share',
'https://www.dianping.com/shopshare/k5NAblkRINRIs3yK?msource=Appshare2021&utm_source=shop_share',
'https://www.dianping.com/shopshare/l2sxaMMccubJImlo?msource=Appshare2021&utm_source=shop_share',
'https://www.dianping.com/shopshare/k4UhphPhTxE2z0zR?msource=Appshare2021&utm_source=shop_share',
'https://www.dianping.com/shopshare/G1nMOSKO1x85gvg4?msource=Appshare2021&utm_source=shop_share',
'https://www.dianping.com/shopshare/H7xTJhKzOsYZbOLX?msource=Appshare2021&utm_source=shop_share',
'https://www.dianping.com/shopshare/E8ZuisuglYHK3iMc?msource=Appshare2021&utm_source=shop_share',
'https://www.dianping.com/shopshare/l1vx5kArj5b7rZcQ?msource=Appshare2021&utm_source=shop_share',
'https://www.dianping.com/shopshare/H1hbLJDF6vvMe1OP?msource=Appshare2021&utm_source=shop_share',
'https://www.dianping.com/shopshare/k4T7umyp7odOnOjm?msource=Appshare2021&utm_source=shop_share',
'https://www.dianping.com/shopshare/H2jXHXRFYnX3jpuW?msource=Appshare2021&utm_source=shop_share',
'https://www.dianping.com/shopshare/k6u01mieZrBs86k0?msource=Appshare2021&utm_source=shop_share',
'https://www.dianping.com/shopshare/k6SWV2OVZM7NiqPp?msource=Appshare2021&utm_source=shop_share',
'https://www.dianping.com/shopshare/l4zEE4MmQjTMjYhy?msource=Appshare2021&utm_source=shop_share',
'https://www.dianping.com/shopshare/isPNIQEFm4p2gLl3?msource=Appshare2021&utm_source=shop_share',
'https://www.dianping.com/shopshare/l8wOKr7nPxGCMWdE?msource=Appshare2021&utm_source=shop_share',
'https://www.dianping.com/shopshare/G8J3UMeucqUqjyGD?msource=Appshare2021&utm_source=shop_share',
'https://www.dianping.com/shopshare/ErhOBkAbIaEPDS3G?msource=Appshare2021&utm_source=shop_share',

]

while i < len(x):
    webdriver.get(x[i])
    wsd['A'+str(dp)]= x[i]
    wsd['G'+str(dp)]= today
    for j in range(3): # adjust integer value for need
        webdriver.execute_script("window.scrollBy(0, 250)")
        sleep (0.5)
        
    point = webdriver.find_elements(by=By.CSS_SELECTOR, value='div.star-score.wx-view')
    for text in point:
                text = text.text
                print (text)
                wsd['H'+str(dp)]=text
    
    comment = webdriver.find_elements(by=By.CSS_SELECTOR, value='div.reviews.wx-view')
    for text in comment:
        text = text.text
        result = re.sub('条', '', text)  
        print (result)
        wsd['C'+str(dp)]=result
    
    comment = webdriver.find_elements(by=By.CSS_SELECTOR, value='span.reviews.wx-text')
    for text in comment:
        text = text.text
        result = re.sub('条', '', text)  
        print (result)
        wsd['C'+str(dp)]=result
    
    price = webdriver.find_elements(by=By.CSS_SELECTOR, value='div.price.wx-view')
    for text in price:
        text = text.text
        result = re.sub('¥|/人', '', text)  
        print (result)
        wsd['B'+str(dp)]=result
    
    price = webdriver.find_elements(by=By.CSS_SELECTOR, value='div.smallPanel.wx-view>span.price.wx-text')
    for text in price:
        text = text.text
        result = re.sub('¥|/人', '', text)  
        print (result)
        wsd['B'+str(dp)]=result
    
    score = webdriver.find_elements(by=By.CSS_SELECTOR, value='div.scoreText.wx-view')
    for text in score:
        text = text.text
        result = re.split('\s',text)
        column = ['D','E','F']
        c=0
        for s in result:
            r = re.sub('口味:|环境:|服务:|产品:','',s)
            print (r)
            wsd[column[c]+str(dp)]=r
            c+=1
    
    score = webdriver.find_elements(by=By.CSS_SELECTOR, value='span.scoreText.wx-text')
    for text in score:
        text = text.text
        result = re.split('\s',text)
        column = ['D','E','F']
        c=0
        for s in result:
            r = re.sub('口味:|环境:|服务:','',s)
            print (r)
            wsd[column[c]+str(dp)]=r
            c+=1
    dp+=1
    
    Translate = BaiduTranslate('zh','en')
    rank = webdriver.find_elements(by=By.CSS_SELECTOR, value='div.rank-text.wx-view')
    for text in rank:
                text = text.text
                print (text)
                # 只获取一次，并睡眠1秒
                C1 = Translate.BdTrans(text)
                print(C1)
                sleep(1)
                if not C1:
                    sleep(1)
                    # 重试一次翻译
                    C1 = Translate.BdTrans(text)
                wsr['A'+str(dr)]=x[i]
                wsr['B'+str(dr)]=text
                wsr['C'+str(dr)]=str(C1)
                wsr['D'+str(dr)]=today
                dr+=1
               
    
    food = webdriver.find_elements(by=By.CSS_SELECTOR, value='div.dishNameContainer.wx-view>div')
    recommend = webdriver.find_elements(by=By.CSS_SELECTOR, value='div.recommendInfo.wx-view')
    f=0
    while f<len(recommend):
                text = food[f].text
                like = recommend[f].text
                result = re.sub('人推荐', '', like)  
                print (text)
                # 只获取一次，并睡眠1秒
                C2 = Translate.BdTrans(text)
                print(C2)
                sleep(1)
                if not C2:
                    sleep(1)
                    # 重试一次翻译
                    C = Translate.BdTrans(text)
                print (result)
                wsf['A'+str(df)]=x[i]
                wsf['B'+str(df)]=text
                wsf['C'+str(df)]=str(C2)
                wsf['D'+str(df)]=result
                wsf['E'+str(df)]=today
                f+=1
                df+=1
    
    tag = webdriver.find_elements(by=By.CSS_SELECTOR, value='div.tags.shop-tags.wx-view>div')
   
    for item in tag:
                keyword = item.get_attribute('data-keyword')
                like = item.get_attribute('data-hit')
                result = re.sub('"', '', keyword)
                print (result)
                # 只获取一次，并睡眠1秒
                C3 = Translate.BdTrans(result)
                print(C3)
                sleep(1)
                if not C3:
                    sleep(1)
                    # 重试一次翻译
                    C = Translate.BdTrans(result)
                print (like)
                wst['A'+str(dt)]=x[i]
                wst['B'+str(dt)]=result
                wst['C'+str(dt)]=str(C3)
                wst['D'+str(dt)]=like
                wst['E'+str(dt)]=today
                
                dt+=1

    
    i+=1
wb.save(filename = excelname)    
'''
webdriver.quit()

content = MIMEMultipart()  #建立MIMEMultipart物件
content["subject"] = "Outlet List"  #郵件標題
content["from"] = "tseminho@gmail.com"  #寄件者
content["to"] = "minho.tse@wynnpalace.com" #收件者
content.attach(MIMEText("Find attachment"))  #郵件內容


# open the file to be sent 
filename = "dianping.xlsx"
attachment = open(excelname,"rb")
  
# instance of MIMEBase and named as p
p = MIMEBase('application', 'octet-stream')
  
# To change the payload into encoded form
p.set_payload((attachment).read())
  
# encode into base64
encoders.encode_base64(p)
   
p.add_header('Content-Disposition', "attachment; filename= %s" % filename)
  
# attach the instance 'p' to instance 'msg'
content.attach(p)

with smtplib.SMTP(host="smtp.gmail.com", port="587") as smtp:  # 設定SMTP伺服器
    try:
        smtp.ehlo()  # 驗證SMTP伺服器
        smtp.starttls()  # 建立加密傳輸
        smtp.login("tseminho@gmail.com", "nnvfoszedywiiibh")  # 登入寄件者gmail
        smtp.send_message(content)  # 寄送郵件
        print("Complete!")
    except Exception as e:
        print("Error message: ", e)
'''